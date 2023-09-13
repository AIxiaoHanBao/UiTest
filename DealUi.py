import sys

from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QApplication, QMainWindow, QPushButton, QVBoxLayout, QWidget, QFileDialog, QLabel, QListWidget, QListWidgetItem, QCheckBox, QMessageBox
import ShowUntil as showUntil
class ExcelConfigPage(QWidget):
    def __init__(self):
        super().__init__()

        self.excel_file_path = None
        self.selected_sheets = []

        self.initUI()

    def initUI(self):
        layout = QVBoxLayout()

        self.excel_button = QPushButton("浏览 Excel 文件")
        self.excel_button.clicked.connect(self.browseExcelFile)
        layout.addWidget(self.excel_button)

        self.excel_label = QLabel("Excel 文件中的表名：")
        layout.addWidget(self.excel_label)

        self.excel_list = QListWidget()
        layout.addWidget(self.excel_list)

        self.select_all_checkbox = QCheckBox("全选")
        self.select_all_checkbox.stateChanged.connect(self.selectAllSheets)
        layout.addWidget(self.select_all_checkbox)

        self.confirm_button = QPushButton("确定")
        self.confirm_button.clicked.connect(self.confirmSelection)
        layout.addWidget(self.confirm_button)

        self.setLayout(layout)

    def browseExcelFile(self):
        options = QFileDialog.Options()
        file_name, _ = QFileDialog.getOpenFileName(self, "选择 Excel 文件", "", "Excel Files (*.xlsx *.xls);;All Files (*)", options=options)
        if file_name:
            self.excel_file_path = file_name
            self.loadExcelSheets(file_name)

    def loadExcelSheets(self, file_path):
        from openpyxl import load_workbook
        self.selected_sheets = showUntil.getExclSheet(r"F:\咸鱼项目\8-8咸鱼unet项目\测试\testAccess\newTest\字段表演示.xlsx")

        self.excel_list.clear()
        self.selected_sheets = []

        try:
            workbook = load_workbook(file_path, read_only=True)
            sheet_names = workbook.sheetnames

            for sheet_name in sheet_names:
                item = QListWidgetItem(sheet_name)
                item.setFlags(item.flags() | Qt.ItemIsUserCheckable)
                item.setCheckState(Qt.Unchecked)
                self.excel_list.addItem(item)

        except Exception as e:
            QMessageBox.warning(self, "错误", f"无法读取 Excel 文件：{str(e)}")

    def selectAllSheets(self, state):
        for i in range(self.excel_list.count()):
            item = self.excel_list.item(i)
            if state == Qt.Checked:
                item.setCheckState(Qt.Checked)
            else:
                item.setCheckState(Qt.Unchecked)

    def confirmSelection(self):
        self.selected_sheets = [item.text() for item in self.excel_list.selectedItems()]
        if not self.selected_sheets:
            QMessageBox.warning(self, "警告", "请至少选择一个 Excel 表")
        else:
            QMessageBox.information(self, "信息", f"已选择的表：{', '.join(self.selected_sheets)}")

def main():
    app = QApplication(sys.argv)
    window = QMainWindow()
    page = ExcelConfigPage()
    window.setCentralWidget(page)
    window.setWindowTitle("Excel 表配置")
    window.show()
    sys.exit(app.exec_())

if __name__ == '__main__':
    main()
